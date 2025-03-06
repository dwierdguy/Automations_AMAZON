# Script to process workflow data and generate Excel reports
# Handles both Albacore and Paramount workflow types
 
import json
import os
import openpyxl
from pathlib import Path
import openpyxl as xl
import getpass
from openpyxl.styles import PatternFill
import time
from . import views

# Global variables for tracking workflow components and state
c_question_id_bucket = []       # Stores question IDs
c_question_group_bucket = []    # Stores question group IDs
c_answer_id_bucket = []         # Stores answer IDs
data_store_dict = {}            # Stores workflow data
workflow_ids = []               # Stores workflow IDs
progress = 0                    # Tracks processing progress
number_of_workflows = 0         # Total number of workflows


def color_fixer(x, y, code):
    """
    Applies color formatting to Excel cells
    Args:
        x: Row number
        y: Column number
        code: Color code
    """
    global wb
    global sheet, worksheet_name
    color_fill = sheet.cell(x, y)
    color_fill.fill = PatternFill(fill_type='solid', start_color=code, end_color=code)
    try:
        wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
    except PermissionError:
        print("Error trying to access the file from color_fixer().")


def extracting_mandate_questions():
    """
    Extracts mandatory questions from workflow
    - Processes question groups
    - Handles different question types (radio, checkbox, textbox)
    - Manages follow-up questions
    - Applies formatting to Excel output
    """
    global wb, sheet, sheet_name, question_json, fup_string, worksheet_name
    global mandate_questions, follow_up_list, row_counter, p_question_group_bucket, extracting_mandate_questions_called
    # Function extracting_mandate_questions_called

    if not extracting_mandate_questions_called:
        progress_check()
        extracting_mandate_questions_called = True

    # Question Group ID
    for item in mandate_questions:
        if item not in c_question_group_bucket:
            print("Skipped a mandate which was not found")
            break
        number_of_questions = len(question_json[item]["workflow_questions"])
        for i in range(0, number_of_questions):
            sheet.cell(row_counter, 9).value = item
            color_fixer(row_counter, 9, 'A9D08E')
            question_id = question_json[item]["workflow_questions"][i]["id"]
            sheet.cell(row_counter, 10).value = question_id
            color_fixer(row_counter, 10, 'C6E0B4')

            question_string = question_json[item]["workflow_questions"][i]["question_string"]
            print(question_string)
            sheet.cell(row_counter, 11).value = question_string
            if "answer_eval_attributes" in question_json[item]["workflow_questions"][i]:
                rule_set_name = question_json[item]["workflow_questions"][i]["answer_eval_attributes"]["rule_set_name"]
                print("working for extracting_mandate_questions")
                sheet.cell(row_counter, 16).value = rule_set_name
                color_fixer(row_counter, 16, 'CC66FF')
                schema_name = question_json[item]["workflow_questions"][i]["answer_eval_attributes"]["schema_name"]
                sheet.cell(row_counter, 17).value = schema_name
                color_fixer(row_counter, 17, 'CC66FF')

            if "radio_options" in question_json[item]["workflow_questions"][i]["responses"][0]:
                number_of_answers = len(question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"])
            elif "checkbox_options" in question_json[item]["workflow_questions"][i]["responses"][0]:
                number_of_answers = len(question_json[item]["workflow_questions"][i]["responses"][0]["checkbox_options"])
            else:
                number_of_answers = 1
            for n in range(0, number_of_answers):

                sheet.cell(row_counter, 9).value = item
                color_fixer(row_counter, 9, 'A9D08E')
                sheet.cell(row_counter, 10).value = question_id
                color_fixer(row_counter, 10, 'C6E0B4')
                sheet.cell(row_counter, 11).value = question_string
                color_fixer(row_counter, 11, '9BC2E6')

                if "radio_options" in question_json[item]["workflow_questions"][i]["responses"][0]:
                    sheet.cell(row_counter, 14).value = "radio"
                    color_fixer(row_counter, 14, 'F8CBAD')
                    answer_id = question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]["id"]
                    if answer_id.startswith("Q-"):
                        answer_id = answer_id.replace("Q-", "A-")
                    if answer_id.startswith("q-"):
                        answer_id = answer_id.replace("q-", "A-")
                    sheet.cell(row_counter, 12).value = answer_id
                    color_fixer(row_counter, 12, 'F8CBAD')
                    answer_label = question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]["label"]
                    # print(answer_id)
                    sheet.cell(row_counter, 13).value = answer_label
                    color_fixer(row_counter, 13, 'F8CBAD')
                    if "next_node_override" in question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]:
                        jump = question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]["next_node_override"]
                        sheet.cell(row_counter, 7).value = jump
                        color_fixer(row_counter, 7, 'FFD966')
                    if "auto_annotation" in question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]:
                        auto_annotation = question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]["auto_annotation"]
                        sheet.cell(row_counter, 15).value = auto_annotation
                        color_fixer(row_counter, 15, 'F8CBAD')
                    if "followup_question_group_ids" in question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]:
                        follow_up = question_json[item]["workflow_questions"][i]["responses"][0]["radio_options"][n]["followup_question_group_ids"]
                        for fps in follow_up:
                            fup_string = fup_string + "," + fps
                            if fps not in follow_up_list:
                                follow_up_list.append(fps)
                        fup_string = fup_string.lstrip(",")
                        fup_string = fup_string.rstrip(",")
                        sheet.cell(row_counter, 8).value = fup_string
                        color_fixer(row_counter, 8, 'FFE699')
                        fup_string = ''
                        # print(type(follow_up))

                if "checkbox_options" in question_json[item]["workflow_questions"][i]["responses"][0]:
                    sheet.cell(row_counter, 14).value = "checkbox"
                    color_fixer(row_counter, 14, 'F8CBAD')
                    answer_id = question_json[item]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]["id"]
                    if answer_id.startswith("Q-"):
                        answer_id = answer_id.replace("Q-", "A-")
                    if answer_id.startswith("q-"):
                        answer_id = answer_id.replace("q-", "A-")
                    sheet.cell(row_counter, 12).value = answer_id
                    color_fixer(row_counter, 12, 'F8CBAD')
                    answer_label = question_json[item]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]["label"]
                    sheet.cell(row_counter, 13).value = answer_label
                    color_fixer(row_counter, 13, 'F8CBAD')
                    if "auto_annotation" in question_json[item]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]:
                        auto_annotation = question_json[item]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]["auto_annotation"]
                        sheet.cell(row_counter, 15).value = auto_annotation
                        color_fixer(row_counter, 15, 'F8CBAD')
                if "max_length" in question_json[item]["workflow_questions"][i]["responses"][0]:
                    sheet.cell(row_counter, 14).value = "textbox"
                    color_fixer(row_counter, 14, 'F8CBAD')
                    answer_id = question_json[item]["workflow_questions"][i]["responses"][0]["id"]
                    if answer_id.startswith("Q-"):
                        answer_id = answer_id.replace("Q-", "A-")
                    if answer_id.startswith("q-"):
                        answer_id = answer_id.replace("q-", "A-")
                    sheet.cell(row_counter, 12).value = answer_id
                    color_fixer(row_counter, 12, 'F8CBAD')
                    max_length = question_json[item]["workflow_questions"][i]["responses"][0]["max_length"]
                    sheet.cell(row_counter, 13).value = max_length
                    color_fixer(row_counter, 13, 'F8CBAD')
                row_counter += 1
                #print(row_counter)
        wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
        row_counter += 1
    row_counter -= 1
    # Call a function for follow-ups and then return to fetching the next section
    if len(follow_up_list) > 0:
        extracting_follow_up()
    else:
        row_counter += 2


def extracting_follow_up():
    """
    Processes follow-up questions
    - Similar to extracting_mandate_questions but for follow-up logic
    - Handles conditional question flows
    - Updates Excel with follow-up data
    """
    global wb, sheet, sheet_name, question_json, fup_string, worksheet_name
    global mandate_questions, follow_up_list, row_counter, c_question_group_bucket, extracting_follow_up_called
    # print(follow_up_list)

    if not extracting_follow_up_called:
        progress_check()
        extracting_follow_up_called = True

    row_counter += 1
    for fups in follow_up_list:
        if fups not in c_question_group_bucket:
            row_counter -= 1
            break
        if fups in question_json:
            number_of_questions = len(question_json[fups]["workflow_questions"])
            for i in range(0, number_of_questions):
                sheet.cell(row_counter, 9).value = fups
                color_fixer(row_counter, 9, 'A9D08E')
                question_id = question_json[fups]["workflow_questions"][i]["id"]
                sheet.cell(row_counter, 10).value = question_id
                color_fixer(row_counter, 10, 'C6E0B4')
                if "question_string" in question_json[fups]["workflow_questions"][i]:
                    question_string = question_json[fups]["workflow_questions"][i]["question_string"]
                else:
                    question_string = "--"

                sheet.cell(row_counter, 11).value = question_string
                color_fixer(row_counter, 11, '9BC2E6')

                if "answer_eval_attributes" in question_json[fups]["workflow_questions"][i]:
                    rule_set_name = question_json[fups]["workflow_questions"][i]["answer_eval_attributes"]["rule_set_name"]
                    print("working for extracting_follow_up")
                    sheet.cell(row_counter, 16).value = rule_set_name
                    color_fixer(row_counter, 16, 'CC66FF')
                    schema_name = question_json[fups]["workflow_questions"][i]["answer_eval_attributes"]["schema_name"]
                    sheet.cell(row_counter, 17).value = schema_name
                    color_fixer(row_counter, 17, 'CC66FF')

                if "radio_options" in question_json[fups]["workflow_questions"][i]["responses"][0]:
                    number_of_answers = len(question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"])
                elif "checkbox_options" in question_json[fups]["workflow_questions"][i]["responses"][0]:
                    number_of_answers = len(question_json[fups]["workflow_questions"][i]["responses"][0]["checkbox_options"])
                else:
                    number_of_answers = 1
                for n in range(0, number_of_answers):

                    sheet.cell(row_counter, 9).value = fups
                    color_fixer(row_counter, 9, 'A9D08E')
                    sheet.cell(row_counter, 10).value = question_id
                    color_fixer(row_counter, 10, 'C6E0B4')
                    sheet.cell(row_counter, 11).value = question_string
                    color_fixer(row_counter, 11, '9BC2E6')

                    if "radio_options" in question_json[fups]["workflow_questions"][i]["responses"][0]:
                        sheet.cell(row_counter, 14).value = "radio"
                        color_fixer(row_counter, 14, 'F8CBAD')
                        answer_id = question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]["id"]
                        if answer_id.startswith("Q-"):
                            answer_id = answer_id.replace("Q-", "A-")
                        if answer_id.startswith("q-"):
                            answer_id = answer_id.replace("q-", "A-")
                        sheet.cell(row_counter, 12).value = answer_id
                        color_fixer(row_counter, 12, 'F8CBAD')
                        answer_label = question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]["label"]
                        #print(answer_label)
                        sheet.cell(row_counter, 13).value = answer_label
                        color_fixer(row_counter, 13, 'F8CBAD')
                        if "next_node_override" in question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]:
                            jump = question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]["next_node_override"]
                            sheet.cell(row_counter, 7).value = jump
                            color_fixer(row_counter, 7, 'FFD966')
                        if "auto_annotation" in question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]:
                            auto_annotation = question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]["auto_annotation"]
                            sheet.cell(row_counter, 15).value = auto_annotation
                            color_fixer(row_counter, 15, 'F8CBAD')
                        if "followup_question_group_ids" in question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]:
                            follow_up = question_json[fups]["workflow_questions"][i]["responses"][0]["radio_options"][n]["followup_question_group_ids"]
                            for fps in follow_up:
                                fup_string = fup_string + "," + fps
                                if fps not in follow_up_list:
                                    follow_up_list.append(fps)
                            fup_string = fup_string.lstrip(",")
                            fup_string = fup_string.rstrip(",")
                            sheet.cell(row_counter, 8).value = fup_string
                            color_fixer(row_counter, 8, 'FFE699')
                            fup_string = ''
                    if "checkbox_options" in question_json[fups]["workflow_questions"][i]["responses"][0]:
                        sheet.cell(row_counter, 14).value = "checkbox"
                        color_fixer(row_counter, 14, 'F8CBAD')
                        answer_id = question_json[fups]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]["id"]
                        if answer_id.startswith("Q-"):
                            answer_id = answer_id.replace("Q-", "A-")
                        if answer_id.startswith("q-"):
                            answer_id = answer_id.replace("q-", "A-")
                        sheet.cell(row_counter, 12).value = answer_id
                        color_fixer(row_counter, 12, 'F8CBAD')
                        answer_label = question_json[fups]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]["label"]
                        sheet.cell(row_counter, 13).value = answer_label
                        color_fixer(row_counter, 13, 'F8CBAD')
                        if "auto_annotation" in question_json[fups]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]:
                            auto_annotation = question_json[fups]["workflow_questions"][i]["responses"][0]["checkbox_options"][n]["auto_annotation"]
                            sheet.cell(row_counter, 15).value = auto_annotation
                            color_fixer(row_counter, 15, 'F8CBAD')
                    if "max_length" in question_json[fups]["workflow_questions"][i]["responses"][0]:
                        sheet.cell(row_counter, 14).value = "textbox"
                        color_fixer(row_counter, 14, 'F8CBAD')
                        answer_id = question_json[fups]["workflow_questions"][i]["responses"][0]["id"]
                        if answer_id.startswith("Q-"):
                            answer_id = answer_id.replace("Q-", "A-")
                        if answer_id.startswith("q-"):
                            answer_id = answer_id.replace("q-", "A-")
                        sheet.cell(row_counter, 12).value = answer_id
                        color_fixer(row_counter, 12, 'F8CBAD')
                        max_length = question_json[fups]["workflow_questions"][i]["responses"][0]["max_length"]
                        sheet.cell(row_counter, 13).value = max_length
                        color_fixer(row_counter, 13, 'F8CBAD')

                    row_counter += 1
        wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
        row_counter += 1
    row_counter += 1
    follow_up_list = []


def identify_starting_section():
    """
    Identifies and processes the initial section of workflow
    - Sets up Excel headers
    - Processes widgets and their visibility
    - Handles section transitions
    - Manages workflow navigation logic
    """
    global wb, sheet, sheet_name, worksheet_name
    global mandate_questions, row_counter, workflow_section, workflow_question, identify_starting_section_called
    section_file = data_store_dict[workflow_section]

    if not identify_starting_section_called:
        progress_check()
        identify_starting_section_called = True

    section_json = json.loads(section_file)
    mandate_question_string = ''
    section_id = []
    not_first_section = []
    # first_section = ''
    first_list = ["Workflow ID", "Widget Combination", "Section ID", "Section Name", "NextSect", "Question groups",
                  "Jump", "FUP Tag", "Q group ID", "Q ID", "Question", "Answer ID", "Answer", "Answer Type",
                  "Annotations", "Eval Ruleset", "Eval Schema"]
    counter = 1
    for heads in first_list:
        sheet.cell(1, counter).value = heads
        color_fixer(1, counter, 'FFD966')
        counter += 1

    # *********************************   W I D G E T S   ****************************************
    customer_data = ["triton-jupiter-customer-data", "triton-jupiter-customer-data-shadow", "triton-customer-data"]
    annotations = ["triton-recent-annotation", "triton-annotation-shadow", "triton-annotation"]
    order_details = ["triton-jupiter-order-data", "triton-jupiter-order-data-shadow", "triton-order-data"]
    credit_cards = ["triton-jupiter-customer-credit-cards", "triton-jupiter-customer-credit-cards-shadow", "triton-payment-methods"]
    addresses = ["triton-addresses", "triton-addresses-shadow", "triton-addressesv2"]
    gsi = ["triton-jupiter-customer-signins", "triton-jupiter-customer-signins-shadow"]
    beta_gsi = ["triton-interesting-signins", "triton-interesting-signins-shadow"]
    rc = ["triton-jupiter-related-customers", "triton-jupiter-related-customers-shadow", "triton-related-customers"]
    gc = ["triton-gift-cards",	"triton-gift-cards-shadow", "triton-gift-cardsv2"]
    # *********************************************************************************************
    sheet.cell(row_counter, 1).value = workflow_id
    for name, properties in section_json.items():
        section_id.append(name)
    sections_list = []
    sections_list = sections_list.append(data_store_dict[workflow_id])
    first_section = []
    sections_list_number = []

    # ****** SET UP COUNT ******
    f_section = data_store_dict[workflow_id]
    sections_list_number.append(f_section)
    while True:
        if "next_workflow_node" in section_json[f_section]:
            next_section = section_json[f_section]["next_workflow_node"]
            sections_list_number.append(next_section)
            f_section = next_section
        else:
            break
    # **************************

    for name, properties in section_json.items():
        if "next_workflow_node" in properties:
            next_section = properties["next_workflow_node"]
            not_first_section.append(next_section)
    for item in section_id:
        if item not in not_first_section:
            first_section.append(item)
    minus_length = len(first_section) - 1
    number_of_sections = len(sections_list_number)
    print(f"Number of sections : {number_of_sections}")

    sheet.cell(row_counter, 3).value = data_store_dict[workflow_id]
    color_fixer(row_counter, 3, 'FFD966')
    section_id = data_store_dict[workflow_id]
    section_name = section_json[section_id]["section_name"]
    sheet.cell(row_counter, 4).value = section_name
    color_fixer(row_counter, 4, 'FFD966')
    next_section_name = section_json[section_id]["next_workflow_node"]
    sheet.cell(row_counter, 5).value = next_section_name
    color_fixer(row_counter, 5, 'FFD966')
    number_of_mandates = len(section_json[section_id]["workflow_question_group_ids"])
    widgets = []
    widget_text = ''
    widgets_id = section_json[section_id]["widgets_visible"]
    # print(widgets_id)
    for wds in widgets_id:
        if wds in customer_data:
            if "Customer data" not in widgets:
                widgets.append("Customer data")
        if wds in annotations:
            if "Annotations" not in widgets:
                widgets.append("Annotations")
        if wds in order_details:
            if "Order Details" not in widgets:
                widgets.append("Order Details")
        if wds in credit_cards:
            if "Credit Cards" not in widgets:
                widgets.append("Credit Cards")
        if wds in addresses:
            if "Addresses" not in widgets:
                widgets.append("Addresses")
        if wds in gsi:
            if "GSI" not in widgets:
                widgets.append("GSI")
        if wds in beta_gsi:
            if "Beta GSI" not in widgets:
                widgets.append("Beta GSI")
        if wds in rc:
            if "RC" not in widgets:
                widgets.append("RC")
        if wds in gc:
            if "GC" not in widgets:
                widgets.append("GC")

    for x in widgets:
        widget_text = widget_text + x + ','
    widget_text = widget_text.lstrip(',')
    widgets_id = []
    widget_text = widget_text.rstrip(',')
    widgets = []
    sheet.cell(row_counter, 2).value = widget_text
    color_fixer(row_counter, 2, 'BF8F00')
    widget_text = ''

    for i in range(0, number_of_mandates):
        mandate_question = section_json[section_id]["workflow_question_group_ids"][i]
        mandate_question_string = mandate_question_string + ',' + mandate_question
        mandate_questions.append(mandate_question)
    mandate_question_string = mandate_question_string.lstrip(',')
    mandate_question_string = mandate_question_string.rstrip(',')
    sheet.cell(row_counter, 6).value = mandate_question_string
    color_fixer(row_counter, 6, 'FFD966')
    mandate_question_string = ''
    # Func() call below works for the first section.
    extracting_mandate_questions()
    mandate_questions = []
    # Working code for sections except for first section
    next_section_id = section_json[section_id]["next_workflow_node"]
    for t in range(0, number_of_sections):
        if t == number_of_sections-2:
            print(f"Entering the last section with 't' value - {t}")
            last_section_id = "node-summary"
            row_counter -= 1
            sheet.cell(row_counter, 3).value = last_section_id
            color_fixer(row_counter, 3, 'FFD966')
            widgets = []
            widget_text = ''
            widgets_id = section_json[last_section_id]["widgets_visible"]
            for wds in widgets_id:
                if wds in customer_data:
                    if "Customer data" not in widgets:
                        widgets.append("Customer data")
                if wds in annotations:
                    if "Annotations" not in widgets:
                        widgets.append("Annotations")
                if wds in order_details:
                    if "Order Details" not in widgets:
                        widgets.append("Order Details")
                if wds in credit_cards:
                    if "Credit Cards" not in widgets:
                        widgets.append("Credit Cards")
                if wds in addresses:
                    if "Addresses" not in widgets:
                        widgets.append("Addresses")
                if wds in gsi:
                    if "GSI" not in widgets:
                        widgets.append("GSI")
                if wds in beta_gsi:
                    if "Beta GSI" not in widgets:
                        widgets.append("Beta GSI")
                if wds in rc:
                    if "RC" not in widgets:
                        widgets.append("RC")
                if wds in gc:
                    if "GC" not in widgets:
                        widgets.append("GC")

            for x in widgets:
                widget_text = widget_text + x + ','
            widget_text = widget_text.lstrip(',')
            widgets_id = []
            widget_text = widget_text.rstrip(',')
            widgets = []
            sheet.cell(row_counter, 2).value = widget_text
            color_fixer(row_counter, 2, 'BF8F00')
            widget_text = ''
            wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
            wb.close()
            other_sheets()
            print("Working")
            toc = time.perf_counter()
            print(f"Mastersheet completed in {toc - tic:0.4f} seconds")
            # exit(1)
            return
            # Keeping it for the last section which does not point to any other section.
        row_counter -= 1
        current_section_id = next_section_id
        print(t, current_section_id)
        sheet.cell(row_counter, 3).value = current_section_id
        color_fixer(row_counter, 3, 'FFD966')
        print(current_section_id)
        if "section_name" in section_json[current_section_id]:
            current_section_name = section_json[current_section_id]["section_name"]
            sheet.cell(row_counter, 4).value = current_section_name
            color_fixer(row_counter, 4, 'FFD966')
        if "workflow_question_group_ids" in section_json[current_section_id]:
            number_of_mandates = len(section_json[current_section_id]["workflow_question_group_ids"])
        else:
            number_of_mandates = 0

        widgets = []
        widget_text = ''
        widgets_id = section_json[current_section_id]["widgets_visible"]
        # print(widgets_id)
        for wds in widgets_id:
            if wds in customer_data:
                if "Customer data" not in widgets:
                    widgets.append("Customer data")
            if wds in annotations:
                if "Annotations" not in widgets:
                    widgets.append("Annotations")
            if wds in order_details:
                if "Order Details" not in widgets:
                    widgets.append("Order Details")
            if wds in credit_cards:
                if "Credit Cards" not in widgets:
                    widgets.append("Credit Cards")
            if wds in addresses:
                if "Addresses" not in widgets:
                    widgets.append("Addresses")
            if wds in gsi:
                if "GSI" not in widgets:
                    widgets.append("GSI")
            if wds in beta_gsi:
                if "Beta GSI" not in widgets:
                    widgets.append("Beta GSI")
            if wds in rc:
                if "RC" not in widgets:
                    widgets.append("RC")
            if wds in gc:
                if "GC" not in widgets:
                    widgets.append("GC")

        for x in widgets:
            widget_text = widget_text + x + ','
        widget_text = widget_text.lstrip(',')
        widgets_id = []
        widget_text = widget_text.rstrip(',')
        widgets = []
        sheet.cell(row_counter, 2).value = widget_text
        color_fixer(row_counter, 2, 'BF8F00')
        widget_text = ''
        if number_of_mandates > 0:
            for i in range(0, number_of_mandates):
                mandate_question = section_json[current_section_id]["workflow_question_group_ids"][i]
                mandate_question_string = mandate_question_string + ',' + mandate_question
                mandate_questions.append(mandate_question)
            mandate_question_string = mandate_question_string.lstrip(',')
            mandate_question_string = mandate_question_string.rstrip(',')
            sheet.cell(row_counter, 6).value = mandate_question_string
            color_fixer(row_counter, 6, 'FFD966')
            mandate_question_string = ''
            next_section_id = section_json[current_section_id]["next_workflow_node"]
            sheet.cell(row_counter, 5).value = next_section_id
            color_fixer(row_counter, 5, 'FFD966')
            extracting_mandate_questions()
            mandate_questions = []
        else:
            wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
        # row_counter += 1
    wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)


def current_question_data():
    """
    Validates and analyzes current question data
    - Checks for duplicate answer IDs
    - Validates question-answer relationships
    - Verifies jump logic
    - Creates validation report
    """
    global workflow_question, workflow_section
    f_name = 'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + workflow_id + "_validator.txt"
    output = Path(f_name)
    if output.is_file():
        print("Deleting Previous Output txt file.")
        os.remove(f'{f_name}')
        os.system(f'type nul > {f_name}')
    c_question_group_bucket_duplicate = []
    duplicated_answer_ids = []
    sections_data = data_store_dict[workflow_section]
    c_section = json.loads(data_store_dict[workflow_section])
    sections_id_bucket = []
    mandates = []
    auto_answer = []
    jumps_used = []
    wrong_jumps = []
    mandate_not_found = []
    for values, data in c_section.items():
        sections_id_bucket.append(values)
        if "workflow_question_group_ids" in data:
            number_of_mandates = len(data["workflow_question_group_ids"])
            for n in range(0, number_of_mandates):
                mandates.append(data['workflow_question_group_ids'][n])

    # ---------------*************----------------
    questions_data = data_store_dict[workflow_question]
    c_questions = json.loads(data_store_dict[workflow_question])
    follow_up_bucket = []
    q_groups_with_no_parent = []
    follow_ups = ''
    c_answer_ids = ''
    global c_question_id_bucket
    global c_question_group_bucket
    global c_answer_id_bucket
    c_length_of_answers = 0
    for c_group_id, c_data in c_questions.items():
        c_question_group_bucket.append(c_group_id)
        c_length_of_questions = len(c_data["workflow_questions"])
        for x in range(0, c_length_of_questions):
            c_question_group_counter = c_group_id
            c_counter = 0
            if "radio_options" in c_data["workflow_questions"][x]["responses"][0]:
                c_length_of_answers = len(c_data["workflow_questions"][x]["responses"][0]["radio_options"])
            elif "checkbox_options" in c_data["workflow_questions"][x]["responses"][0]:
                c_length_of_answers = len(c_data["workflow_questions"][x]["responses"][0]["checkbox_options"])
            else:
                c_length_of_answers = 1
            c_question_ids = c_data["workflow_questions"][x]["id"]
            c_question_id_bucket.append(c_question_ids)
            for i in range(0, c_length_of_answers):
                if "radio_options" in c_data["workflow_questions"][x]["responses"][0]:
                    c_answer_ids = c_data["workflow_questions"][x]["responses"][0]["radio_options"][c_counter]["id"]
                    if "followup_question_group_ids" in c_data["workflow_questions"][x]["responses"][0]["radio_options"][c_counter]:
                        length_of_followups = len(c_data["workflow_questions"][x]["responses"][0]["radio_options"][c_counter]["followup_question_group_ids"])
                        for t in range(0, length_of_followups):
                            follow_ups = c_data["workflow_questions"][x]["responses"][0]["radio_options"][c_counter]["followup_question_group_ids"][t]
                            follow_up_bucket.append(follow_ups)
                    if "next_node_override" in c_data["workflow_questions"][x]["responses"][0]["radio_options"][c_counter]:
                        jump = c_data["workflow_questions"][x]["responses"][0]["radio_options"][c_counter]["next_node_override"]
                        jumps_used.append(c_answer_ids)
                        if jump not in sections_id_bucket:
                            wrong_jumps.append(c_answer_ids)
                elif "checkbox_options" in c_data["workflow_questions"][x]["responses"][0]:
                    c_answer_ids = c_data["workflow_questions"][x]["responses"][0]["checkbox_options"][c_counter]["id"]
                else:
                    c_answer_ids = c_data["workflow_questions"][x]["responses"][0]["id"]
                if c_answer_ids not in c_answer_id_bucket:
                    c_answer_id_bucket.append(c_answer_ids)
                else:
                    duplicated_answer_ids.append(c_answer_ids)
                if c_question_group_counter == c_group_id:
                    c_counter += 1
                else:
                    c_counter = 0
                i += 1
            if "answer_eval_attributes" in c_data["workflow_questions"][x]:
                auto_answer.append(c_group_id)
            x += 1
    duplicate_list = follow_up_bucket
    follow_up_bucket = []
    for items in duplicate_list:
        if items not in follow_up_bucket and items != '':
            follow_up_bucket.append(items)

    for items in follow_up_bucket:
        if items not in c_question_group_bucket:
            print("\nFollow up " + items + " not linked to it's parent question.\n", file=open(f_name, "a"))

    # Removing the mandates from the question group bucket

    c_question_group_bucket_duplicate.extend(c_question_group_bucket)
    for item in mandates:
        if item in c_question_group_bucket_duplicate:
            c_question_group_bucket_duplicate.remove(item)

    for items in c_question_group_bucket_duplicate:
        if items not in follow_up_bucket:
            q_groups_with_no_parent.append(items)
    if len(jumps_used) > 0:
        print("\nAnswers using Jumps - ", jumps_used, file=open(f_name, "a"))

    if len(wrong_jumps) > 0:
        print("\nAnswers using wrong jumps - ", wrong_jumps, file=open(f_name, "a"))

    if len(q_groups_with_no_parent) > 0:
        print("\nInactive question groups - ", q_groups_with_no_parent, file=open(f_name, "a"))

    # Checking if mandates from sections.json are present in the questions.json file.

    for items in mandates:
        if items not in c_question_group_bucket:
            mandate_not_found.append(items)

    if len(mandate_not_found) > 0:
        print("Mandate(s) not found in the questions json - ", mandate_not_found, file=open(f_name, "a"))

    if len(duplicated_answer_ids) > 0:
        print("Duplicate answer ids - ", duplicated_answer_ids, file=open(f_name, "a"))

    wrong_answer_ids = []
    for answers in c_answer_id_bucket:
        if answers.startswith("Q-"):
            wrong_answer_ids.append(answers)
    if len(wrong_answer_ids) > 0:
        print("Answer ID that starts with Q- ", wrong_answer_ids, file=open(f_name, "a"))
    if len(auto_answer) > 0:
        print("Questions using auto-answers : ", auto_answer, file=open(f_name, "a"))


def other_sheets():
    """
    Creates additional Excel sheets
    - Copies template data from Sample.xlsx
    - Creates sheets for widgets, answer types, and info
    - Applies formatting and cell merging
    """
    global worksheet_name
    source_file = xl.load_workbook('Sample.xlsm')
    source_sheet1 = source_file.worksheets[0]
    source_sheet2 = source_file.worksheets[1]
    source_sheet3 = source_file.worksheets[2]
    destination_file = xl.load_workbook('C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
    final_sheet1 = destination_file.create_sheet('widget ids')
    final_sheet2 = destination_file.create_sheet('answer_type')
    final_sheet3 = destination_file.create_sheet('info')

    mr = source_sheet1.max_row
    mc = source_sheet1.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = source_sheet1.cell(row=i, column=j)
            final_sheet1.cell(row=i, column=j).value = c.value

    mr = source_sheet2.max_row
    mc = source_sheet2.max_column
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = source_sheet2.cell(row=i, column=j)
            final_sheet2.cell(row=i, column=j).value = c.value

    mr = source_sheet3.max_row
    mc = source_sheet3.max_column
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = source_sheet3.cell(row=i, column=j)
            final_sheet3.cell(row=i, column=j).value = c.value

    # saving the destination excel file
    final_sheet3.merge_cells('A1:C2')
    final_sheet3.merge_cells('D1:F2')
    final_sheet3.merge_cells('A3:C4')
    final_sheet3.merge_cells('D3:F4')
    destination_file.save('C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)


def web_automated_data(data_dict, wf_ids):
    """
    Main function to process workflow data
    Args:
        data_dict: Dictionary containing workflow data
        wf_ids: List of workflow IDs to process
    - Creates output directory if needed
    - Processes each workflow
    - Handles exceptions for unstable workflows
    - Manages progress tracking
    """
    current_f = ''
    global data_store_dict, workflow_ids, tic, mandate_questions, \
        follow_up_list, wb, sheet, sheet_name, row_counter, section_file, \
        question_json, fup_string, worksheet_name, workflow_id, \
        workflow_section, workflow_question, extracting_mandate_questions_called, \
        extracting_follow_up_called, identify_starting_section_called
    global workflow_ids
    data_store_dict = data_dict
    workflow_ids = wf_ids

    # making a folder 'Source files' in Desktop if not already created
    path_dir = 'C:\\Users\\' + getpass.getuser() + '\Desktop\\Source files'
    check_folder = os.path.isdir(path_dir)
    if not check_folder:
        print("Not present")
        path = os.path.join(path_dir)
        os.mkdir(path)
        print("New directory created!")
    else:
        print("Folder already present")

    for w_id in workflow_ids:
        try:
            web_automated_data_called = False
            extracting_mandate_questions_called = False
            extracting_follow_up_called = False
            identify_starting_section_called = False
            current_f = w_id
            if not web_automated_data_called:
                progress_check()
                web_automated_data_called = True
            workflow_id = w_id
            worksheet_name = workflow_id + ".xlsx"
            tic = time.perf_counter()
            mandate_questions = []
            follow_up_list = []
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "master sheet"
            sheet_name = sheet.title
            wb.save(f'C:\\Users\\'+getpass.getuser()+'\Desktop\\Source files\\' + worksheet_name)
            print("Parser collected the workflows IDs - ", workflow_id)
            row_counter = 2
            workflow_section = "sections_" + workflow_id
            workflow_question = "questions_" + workflow_id
            question_file = data_store_dict[workflow_question]
            section_file = data_store_dict[workflow_section]
            question_json = json.loads(question_file)
            fup_string = ''

            current_question_data()

            identify_starting_section()

        except Exception as e:
            print("Workflow unstable : ", current_f)
            print("Error noted : ", e)
            continue


def progress_check():
    """
    Updates progress tracking
    - Increments progress counter
    - Tracks total number of workflows
    """
    global progress, workflow_ids
    progress += 0.25
    number_of_workflows = len(workflow_ids)
    
    
"""
The script's main functions:
    Reads workflow data (questions, sections, answers)
    Processes different question types (mandatory, follow-up, conditional)
    Validates workflow logic and relationships
    Generates formatted Excel reports
    Creates validation reports for workflow issues
    Handles multiple workflows in batch processing
    Manages error handling and progress tracking

Key features:
    Excel report generation with formatted output
    Workflow validation and error checking
    Support for different question types
    Follow-up question handling
    Widget visibility management
    Progress tracking
    Error handling for unstable workflows

The output includes:
    Formatted Excel workbook with multiple sheets
    Validation reports for workflow issues
    Color-coded formatting for better readability
    Organized structure of workflow components
"""
