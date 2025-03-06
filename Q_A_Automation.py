# This code is for extracting Questions and Answers from the BPMN Code. 
# It should extract section name, question variable, question text, answer variable, answer text

from openpyxl import Workbook
from openpyxl import load_workbook
import re
from openpyxl.styles import Alignment

# Initialize a new Excel workbook for storing Q&A data
wb = Workbook()
ws = wb.active
wb.save("QA_collect.xlsx")
wb.close()

# Reopen the workbook and create column headers
wb = load_workbook("QA_collect.xlsx")
ws = wb["Sheet"]

ws.cell(1, 1).value = "Section Name"
ws.cell(1, 2).value = "Question variable"
ws.cell(1, 3).value = "Question"
ws.cell(1, 4).value = "Answer variable"
ws.cell(1, 5).value = "Answer"
ws.cell(1, 6).value = "Follow-up Question"
row = 1

# Dictionary to store question-answer mappings
q_a_dictionary = {}


def get_readable_format(followup_dict, qa_dict):
    """
    Converts the follow-up questions and answers into a readable format
    Args:
        followup_dict: Dictionary containing follow-up question relationships
        qa_dict: Dictionary containing all Q&A mappings
    Returns:
        String containing formatted follow-up questions and their corresponding answers
    """
    readable_outputs = {}
    questions_to_delete = []
    
    for question_id, answers in followup_dict.items():
        if question_id in qa_dict:
            question_label = qa_dict[question_id]['label']
            
            if question_label not in readable_outputs:
                readable_outputs[question_label] = []
            
            for answer_value in answers:
                if answer_value in qa_dict[question_id]:
                    answer_label = qa_dict[question_id][answer_value]
                    if answer_label not in readable_outputs[question_label]:
                        readable_outputs[question_label].append(answer_label)
            
            questions_to_delete.append(question_id)
    
    # Delete the processed questions from the dictionary
    for question_id in questions_to_delete:
        del followup_dict[question_id]
    
    # Format the output as a single string
    formatted_output = ""
    for question, answers in readable_outputs.items():
        if formatted_output:  # Add line break if not the first line
            formatted_output += "\n"
        
        # If there are no answers or empty answers, add (Checkbox)
        if not answers or all(answer.strip() == "" for answer in answers):
            formatted_output += f"{question} > (Checkbox)"
        else:
            formatted_output += f"{question} > {' OR '.join(answers)}"
    
    return formatted_output


def remove_duplication(string):
    """
    Removes duplicated content from a string
    Args:
        string: Input string that might contain duplicated content
    Returns:
        String with duplications removed
    """
    half_length = len(string) // 2
    if string[:half_length] == string[half_length:]:
        return string[:half_length]
    return string


def dict_data_gathering():
    """
    Processes the BPMN XML file to gather all questions and answers into a dictionary
    Handles different question types: YES/NO, CHECKBOX, SELECT/RADIO buttons
    """
    global q_a_dictionary
    radio_options = False
    
    with open(file=r'C:\Users\pratuhin\Downloads\Code repo\Code repo\bpmnxml\paramount.xml', encoding='UTF-8', mode='r') as bpmn_code:
        for lines in bpmn_code:
            # Handle Yes/No questions
            if r'<camunda:formField id= \"' in lines and r'type= \"YES_NO_QUESTION\"' in lines:
                radio_options = False
                lines = lines.split('\"')
                question_variable = lines[1].replace('\\', '')
                question = lines[3].replace('\\', '')
                
                q_a_dictionary[question_variable] = {
                    "label": question,
                    "True": "Yes",
                    "False": "No"
                }
            
            if r'<camunda:formField id= \"' in lines and r'type= \"CHECKBOX\"' in lines:
                pass
            
            
            # Handle Select/Radio questions
            if r'<camunda:formField id=' in lines and any(type in lines for type in [
                r'type= \"SELECT_RADIO\"',
                r'type= \"SELECT_ONE\"',
                r'type= \"SELECT_BUTTON\"'
            ]):
                radio_options = True
                lines = lines.split('\"')
                question_variable = lines[1].replace('\\', '')
                question = lines[3].replace('\\', '')
                
                if question_variable not in q_a_dictionary:
                    q_a_dictionary[question_variable] = {"label": question}
            
            # Process answer options for checkbox questions
            if r'<camunda:formField id=' in lines and r'type= \"CHECKBOX\"' in lines:
                lines = lines.split('\"')
                question_variable = lines[1].replace('\\', '')
                question = lines[3].replace('\\', '')
                
                if question_variable not in q_a_dictionary:
                    q_a_dictionary[question_variable] = {"label": question}
                    
            # Process answer options for radio/select questions
            if r'<camunda:property name= \"options.' in lines and radio_options:
                lines = lines.split('\"')
                answer_label = lines[1].replace('options.', '')
                answer_label = answer_label[:-3]  # Remove last 3 characters
                answer_variable = lines[3].replace('\\', '')
                
                q_a_dictionary[question_variable][answer_variable] = answer_label


# Main processing loop
dict_data_gathering()

# Initialize flags for different question types
radio_options = False
if_checkbox = False
if_textbox = False
yes_no_identifier = False

# Process the BPMN XML file to extract and organize Q&A data
with open(file=r'C:\Users\pratuhin\Downloads\Code repo\Code repo\bpmnxml\paramount.xml', encoding='UTF-8', mode='r') as bpmn_code:
    row += 1
    for lines in bpmn_code:
        if r'<bpmn:userTask id= \"' in lines:
            # To get the section name from the code
            lines = lines.split('\"')
            line = lines[3]
            line = line.replace('\\', '')
            section_name = str(line)
            section_name = section_name.replace('&#10;', '')
            section_name = section_name.replace('&#38', '')
            ws.cell(row, 1).value = section_name
            
            # Restarting the dicrtionary for fresh Q-A stack
            followup_question_dictionary = {}
        
        if r'<camunda:formField id= \"' in lines and r'type= \"YES_NO_QUESTION\"' in lines:
            # To get the question, question variable from the code
            lines = lines.split('\"')
            question_variable = lines[1]
            question_variable = question_variable.replace('\\', '')
            question = lines[3]
            question = question.replace('\\', '')
            ws.cell(row, 2).value = question_variable
            ws.cell(row, 3).value = question
            ws.cell(row, 4).value = "True"
            ws.cell(row, 5).value = "Yes"
            row += 1
            ws.cell(row, 4).value = "False"
            ws.cell(row, 5).value = "No"
            row -= 1
            yes_no_identifier = True

        # Getting the follow-up evaluation condition to save in dictionary
        if r"conditionalShowExpression" in lines and r'eval(' in lines:
            text = lines
            replacements = {"&#39": "", '"': "", ")": "", "(": "", "/>": "", "===": "", ";": "", " ": "", "\\": "", "&#34": ""}
            
            # Create a new question_answer_dict for each line
            question_answer_dict = {}
            
            # Split the text by 'eval(' to get all evaluation expressions
            eval_expressions = text.split('eval(')[1:]
                        
            for expression in eval_expressions:
                # Split each expression into question and answer parts
                parts = expression.split("===")
                if len(parts) == 2:
                    question_var = parts[0].strip()
                    follow_up_answer_var = parts[1].split('||')[0].strip()  # Take only the answer part before any '||'
                    
                    # Apply replacements
                    for old, new in replacements.items():
                        question_var = question_var.replace(old, new)
                        follow_up_answer_var = follow_up_answer_var.replace(old, new)
                    
                    # Remove duplication in question_var
                    question_var = remove_duplication(question_var)
                    
                    # Add to dictionary of lists
                    if question_var not in question_answer_dict:
                        question_answer_dict[question_var] = []
                    question_answer_dict[question_var].append(follow_up_answer_var)
            
            # Update the main dictionary
            for question, answers in question_answer_dict.items():
                if question in followup_question_dictionary:
                    # Add only unique answers
                    followup_question_dictionary[question].extend(
                        [ans for ans in answers if ans not in followup_question_dictionary[question]]
                    )
                else:
                    followup_question_dictionary[question] = answers
                    
            readable_results = get_readable_format(followup_question_dictionary, q_a_dictionary)
    
            if if_checkbox:
                row -= 1
                if_checkbox = False
            # To set the cell formatting for better display
            cell = ws.cell(row, 6)
            cell.value = readable_results
            cell.alignment = Alignment(wrap_text=True)  # Don't forget to import Alignment
                
            if yes_no_identifier:
                row += 2
                yes_no_identifier = False
            
        
        if r'<camunda:formField id=' in lines and r'type= \"CHECKBOX\"' in lines:
            # To get the checkbox value and variable from the code
            lines = lines.split('\"')
            question_variable = lines[1]
            question_variable = question_variable.replace('\\', '')
            question = lines[3]
            question = question.replace('\\', '')
            ws.cell(row, 2).value = question_variable
            ws.cell(row, 3).value = question
            ws.cell(row, 4).value = "true"
            ws.cell(row, 5).value = "Checkbox"
            row += 1
            if_checkbox = True

        if r'<camunda:formField id=' in lines and (r'type= \"SELECT_RADIO\"' in lines or r'type= \"SELECT_ONE\"' in lines or r'type= \"SELECT_BUTTON\"' in lines):
            # To get the question, question variable from the code for select buttons/radio buttons/select box
            if if_checkbox:
                row += 1
                if_checkbox = False
            lines = lines.split('\"')
            question_variable = lines[1]
            question_variable = question_variable.replace('\\', '')
            question = lines[3]
            question = question.replace('\\', '')
            ws.cell(row, 2).value = question_variable
            ws.cell(row, 3).value = question
            radio_options = True

        if r'<camunda:property name= \"options.' in lines and radio_options is True:
            # To get the answers, answer variables from the code
            lines = lines.split('\"')
            answer_label = lines[1]
            answer_label = answer_label.replace('options.', '')
            answer_label = answer_label.rstrip(answer_label[-3:])
            answer_variable = lines[3]
            answer_variable = answer_variable.replace('\\', '')
            ws.cell(row, 4).value = answer_variable
            ws.cell(row, 5).value = answer_label
            row += 1

        if r'<camunda:formField id=' in lines and r'type= \"STRING' in lines:
            lines = lines.split('\"')
            question_variable = lines[1]
            question_variable = question_variable.replace('\\', '')
            question = lines[3]
            question = question.replace('\\', '')
            ws.cell(row, 2).value = question_variable
            ws.cell(row, 3).value = question
            ws.cell(row, 4).value = "Textbox"
            if_textbox = True

        if r'<camunda:constraint name= \"maxBoundary\"' in lines and if_textbox is True:
            lines = lines.split('\"')
            textbox_size = lines[3].replace('\\', '')
            ws.cell(row, 5).value = textbox_size
            row += 2
            if_textbox = False
        elif if_textbox:
            ws.cell(row, 5).value = "Size not assigned"
            row += 2
            if_textbox = False

        if r'</camunda:properties>' in lines and radio_options is True:
            if if_textbox is False:
                row += 1
                radio_options = False
    
    # Save the final Excel workbook
    wb.save('QA_collect.xlsx')

wb.close()




